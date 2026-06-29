import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.chart import BarChart, Reference
import openpyxl

# Load source data
df = pd.read_excel('/home/claude/crypto_dataset.xlsx')

# Load workbook and keep original data sheet
wb = load_workbook('/home/claude/crypto_dataset.xlsx')

# ─── Create Comparison Sheet ────────────────────────────────────────────────
if 'Crypto Comparison' in wb.sheetnames:
    del wb['Crypto Comparison']

ws = wb.create_sheet('Crypto Comparison', 0)  # first sheet

# ─── Style helpers ─────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill('solid', start_color=hex_color, fgColor=hex_color)

def border(style='thin'):
    s = Side(style=style, color='BDBDBD')
    return Border(left=s, right=s, top=s, bottom=s)

def font(bold=False, color='000000', size=11, italic=False):
    return Font(name='Arial', bold=bold, color=color, size=size, italic=italic)

def align(h='center', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

# Color palette
C_DARK_BLUE  = '1A237E'  # header deep navy
C_BLUE       = '1565C0'  # section headers
C_LIGHT_BLUE = 'E3F2FD'  # alternating rows
C_GOLD       = 'F9A825'  # KPI highlight
C_GREEN      = '2E7D32'  # positive accent
C_TEAL       = '00695C'  # coin 2 accent
C_INPUT      = 'FFF9C4'  # input cells
C_WHITE      = 'FFFFFF'
C_GRAY       = 'F5F5F5'
C_DARK_GRAY  = '424242'
C_KPI_BG     = 'E8F5E9'
C_BORDER     = 'BDBDBD'
C_RED        = 'B71C1C'

# ─── Column widths ──────────────────────────────────────────────────────────
ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 26
ws.column_dimensions['C'].width = 30
ws.column_dimensions['D'].width = 3
ws.column_dimensions['E'].width = 30
ws.column_dimensions['F'].width = 3

# ─── Row heights ────────────────────────────────────────────────────────────
for r in range(1, 60):
    ws.row_dimensions[r].height = 22

ws.row_dimensions[1].height = 14
ws.row_dimensions[2].height = 48
ws.row_dimensions[3].height = 14
ws.row_dimensions[4].height = 20
ws.row_dimensions[5].height = 34
ws.row_dimensions[6].height = 14
ws.row_dimensions[7].height = 30
ws.row_dimensions[8].height = 26

# ─── Helper: merge + style cell ─────────────────────────────────────────────
def mset(ws, rng, value='', bg=None, fg='000000', bold=False, sz=11,
         ha='center', va='center', wrap=False, italic=False, num_fmt=None, border_style=None):
    ws.merge_cells(rng)
    start = rng.split(':')[0]
    cell = ws[start]
    cell.value = value
    if bg:
        cell.fill = fill(bg)
    cell.font = Font(name='Arial', bold=bold, color=fg, size=sz, italic=italic)
    cell.alignment = Alignment(horizontal=ha, vertical=va, wrap_text=wrap)
    if num_fmt:
        cell.number_format = num_fmt
    if border_style:
        s = Side(style=border_style, color=C_BORDER)
        cell.border = Border(left=s, right=s, top=s, bottom=s)
    return cell

# ─── TITLE BANNER ───────────────────────────────────────────────────────────
mset(ws, 'B2:E2',
     '🪙  CRYPTO COIN COMPARISON DASHBOARD',
     bg=C_DARK_BLUE, fg=C_WHITE, bold=True, sz=18, ha='center')

# ─── INPUT SECTION HEADER ────────────────────────────────────────────────────
mset(ws, 'B4:E4', '▶  ENTER COIN NAMES TO COMPARE',
     bg=C_BLUE, fg=C_WHITE, bold=True, sz=11)

# ─── Input Label Row ────────────────────────────────────────────────────────
mset(ws, 'B5:B5', 'Metric', bg=C_DARK_BLUE, fg=C_WHITE, bold=True, sz=11)
mset(ws, 'C5:C5', 'Coin Name 1', bg=C_BLUE, fg=C_WHITE, bold=True, sz=12)
mset(ws, 'E5:E5', 'Coin Name 2', bg=C_TEAL, fg=C_WHITE, bold=True, sz=12)

# ─── Coin Name Input Cells (C7 and E7) ──────────────────────────────────────
mset(ws, 'B7:B7', 'Coin Name', bg=C_DARK_BLUE, fg=C_WHITE, bold=True)
ws['C7'].value = 'Bitcoin'
ws['C7'].fill = fill(C_INPUT)
ws['C7'].font = Font(name='Arial', bold=True, color=C_BLUE, size=13)
ws['C7'].alignment = align('center', 'center')
ws['C7'].border = Border(
    left=Side(style='medium', color=C_BLUE),
    right=Side(style='medium', color=C_BLUE),
    top=Side(style='medium', color=C_BLUE),
    bottom=Side(style='medium', color=C_BLUE)
)

ws['E7'].value = 'Ethereum'
ws['E7'].fill = fill(C_INPUT)
ws['E7'].font = Font(name='Arial', bold=True, color=C_TEAL, size=13)
ws['E7'].alignment = align('center', 'center')
ws['E7'].border = Border(
    left=Side(style='medium', color=C_TEAL),
    right=Side(style='medium', color=C_TEAL),
    top=Side(style='medium', color=C_TEAL),
    bottom=Side(style='medium', color=C_TEAL)
)

# ─── VS label ───────────────────────────────────────────────────────────────
ws['D7'].value = 'VS'
ws['D7'].fill = fill(C_GOLD)
ws['D7'].font = Font(name='Arial', bold=True, color='000000', size=12)
ws['D7'].alignment = align('center', 'center')

# ─── DATA VALIDATION on C7 and E7 ───────────────────────────────────────────
# Rule: len > 2 AND len < 11 AND ISNUMBER returns false (no digits)
# Excel custom formula for C7:
dv1 = DataValidation(
    type='custom',
    formula1='=AND(LEN(C7)>2,LEN(C7)<11,SUMPRODUCT(--ISNUMBER(FIND({0,1,2,3,4,5,6,7,8,9},C7)))=0)',
    allow_blank=False,
    showErrorMessage=True,
    errorTitle='Invalid Coin Name',
    error='Coin name must be 3–10 characters long and must not contain numbers.',
    showInputMessage=True,
    promptTitle='Enter Coin Name',
    prompt='Name must be 3–10 characters, letters only (no digits).'
)
dv2 = DataValidation(
    type='custom',
    formula1='=AND(LEN(E7)>2,LEN(E7)<11,SUMPRODUCT(--ISNUMBER(FIND({0,1,2,3,4,5,6,7,8,9},E7)))=0)',
    allow_blank=False,
    showErrorMessage=True,
    errorTitle='Invalid Coin Name',
    error='Coin name must be 3–10 characters long and must not contain numbers.',
    showInputMessage=True,
    promptTitle='Enter Coin Name',
    prompt='Name must be 3–10 characters, letters only (no digits).'
)
dv1.sqref = 'C7'
dv2.sqref = 'E7'
ws.add_data_validation(dv1)
ws.add_data_validation(dv2)

# ─── DETAIL TABLE ───────────────────────────────────────────────────────────
# Rows 9 onward: field rows
mset(ws, 'B9:E9', '📋  COIN SPECIFICATIONS',
     bg=C_BLUE, fg=C_WHITE, bold=True, sz=11)

fields = [
    ('Symbol',             'Symbol',           'B', 11),
    ('Price (USD)',        'Price (USD)',       'C', 12),
    ('Volume (24h)',       'Volume (24h) (USD)','D', 13),
    ('Market Capital',    'Market Cap (USD)',  'E', 14),
    ('Circulating Supply','Circulating Supply', 'F', 15),
]

field_labels = [
    ('Symbol',              10),
    ('Price (USD)',         11),
    ('Volume (24h)',        12),
    ('Market Capital',      13),
    ('Circulating Supply',  14),
]

col_map = {
    'Symbol':             'B',
    'Price (USD)':        'C',
    'Volume (24h) (USD)': 'D',
    'Market Cap (USD)':   'E',
    'Circulating Supply': 'F',
}

# Data sheet name for VLOOKUP
data_sheet = 'Crypto Data'

# Map dataset columns to their positions (1-indexed from A)
# Crypto Data columns: A=Coin Name, B=Symbol, C=Price(USD), ..., N=Market Cap, O=Circulating Supply, P=Volume
# Let's confirm column positions dynamically
col_positions = {col: i+1 for i, col in enumerate(df.columns)}
# A=1(Coin Name), B=2(Symbol), C=3(Price USD), N=14(Market Cap USD), O=15(Circulating Supply), P=16(Volume 24h USD)
sym_col = col_positions['Symbol']           # 2
price_col = col_positions['Price (USD)']    # 3
mktcap_col = col_positions['Market Cap (USD)']  # 14
supply_col = col_positions['Circulating Supply']  # 15
volume_col = col_positions['Volume (24h) (USD)']  # 16

spec_fields = [
    ('Symbol',              sym_col,    None,           'B'),
    ('Price (USD)',         price_col,  '$#,##0.0000',  'C'),
    ('Volume (24h)',        volume_col, '$#,##0',       'D'),
    ('Market Capital',     mktcap_col, '$#,##0',       'E'),
    ('Circulating Supply', supply_col, '#,##0',        'F'),
]

for i, (label, col_idx, num_fmt, _) in enumerate(spec_fields):
    row = 10 + i
    ws.row_dimensions[row].height = 26

    # Label cell
    cell_b = ws.cell(row=row, column=2)
    cell_b.value = label
    cell_b.fill = fill(C_DARK_BLUE if i % 2 == 0 else C_BLUE)
    cell_b.font = Font(name='Arial', bold=True, color=C_WHITE, size=10)
    cell_b.alignment = align('left', 'center')
    cell_b.border = border()

    # Coin 1 VLOOKUP (column C)
    cell_c = ws.cell(row=row, column=3)
    cell_c.value = (
        f"=IFERROR(VLOOKUP($C$7,'Crypto Data'!$A:$R,{col_idx},0),\"Not Found\")"
    )
    cell_c.fill = fill(C_LIGHT_BLUE if i % 2 == 0 else C_WHITE)
    cell_c.font = Font(name='Arial', color='000000', size=11)
    cell_c.alignment = align('center', 'center')
    cell_c.border = border()
    if num_fmt:
        cell_c.number_format = num_fmt

    # Coin 2 VLOOKUP (column E)
    cell_e = ws.cell(row=row, column=5)
    cell_e.value = (
        f"=IFERROR(VLOOKUP($E$7,'Crypto Data'!$A:$R,{col_idx},0),\"Not Found\")"
    )
    cell_e.fill = fill(C_LIGHT_BLUE if i % 2 == 0 else C_WHITE)
    cell_e.font = Font(name='Arial', color='000000', size=11)
    cell_e.alignment = align('center', 'center')
    cell_e.border = border()
    if num_fmt:
        cell_e.number_format = num_fmt

    # D column (separator) - blank styled
    cell_d = ws.cell(row=row, column=4)
    cell_d.fill = fill(C_GOLD)
    cell_d.border = border()
    cell_d.value = '↔'
    cell_d.font = Font(name='Arial', bold=True, color='000000', size=11)
    cell_d.alignment = align('center', 'center')

# ─── KPI SECTION ────────────────────────────────────────────────────────────
kpi_start = 16
ws.row_dimensions[kpi_start].height = 14
ws.row_dimensions[kpi_start+1].height = 20
for r in range(kpi_start+2, kpi_start+10):
    ws.row_dimensions[r].height = 30

mset(ws, f'B{kpi_start}:E{kpi_start}', '',  bg=C_WHITE)
mset(ws, f'B{kpi_start+1}:E{kpi_start+1}',
     '📊  KEY PERFORMANCE INDICATORS  —  DIFFERENCE ANALYSIS',
     bg=C_DARK_BLUE, fg=C_WHITE, bold=True, sz=11)

# KPI definitions: (label, col1_row, col2_row, col_idx, num_fmt, unit)
kpi_defs = [
    ('Volume Difference',            volume_col, '$#,##0',  '24h USD'),
    ('Circulating Supply Difference', supply_col, '#,##0',  'Units'),
    ('Market Capital Difference',    mktcap_col, '$#,##0',  'USD'),
]

kpi_titles = ['📦 Volume Difference', '🔄 Circulating Supply Difference', '💰 Market Capital Difference']
kpi_cols   = [volume_col, supply_col, mktcap_col]
kpi_fmts   = ['$#,##0', '#,##0', '$#,##0']
kpi_colors_bg = ['E3F2FD', 'E8F5E9', 'FFF9C4']
kpi_colors_border = [C_BLUE, C_GREEN, 'F9A825']

# layout: 3 KPIs stacked vertically, each spanning B:E, 3 rows tall
for k in range(3):
    r = kpi_start + 2 + k * 3

    ws.row_dimensions[r].height = 20
    ws.row_dimensions[r+1].height = 36
    ws.row_dimensions[r+2].height = 16

    # Title
    mset(ws, f'B{r}:E{r}', kpi_titles[k],
         bg=kpi_colors_border[k], fg=C_WHITE, bold=True, sz=10)

    # Value cell - difference formula
    # =IFERROR(VLOOKUP(C7,data,col,0) - VLOOKUP(E7,data,col,0), "N/A")
    col_idx = kpi_cols[k]
    diff_formula = (
        f"=IFERROR("
        f"VLOOKUP($C$7,'Crypto Data'!$A:$R,{col_idx},0)"
        f"-VLOOKUP($E$7,'Crypto Data'!$A:$R,{col_idx},0)"
        f",\"N/A\")"
    )
    mset(ws, f'B{r+1}:C{r+1}', diff_formula,
         bg=kpi_colors_bg[k], fg=C_DARK_GRAY, bold=True, sz=16,
         ha='center', num_fmt=kpi_fmts[k], border_style='medium')

    # Context label
    mset(ws, f'D{r+1}:E{r+1}',
         f'Coin1 minus Coin2',
         bg=kpi_colors_bg[k], fg='757575', italic=True, sz=10, ha='center')

    # Spacer
    for c in range(2, 6):
        ws.cell(row=r+2, column=c).fill = fill(C_WHITE)

# ─── Footer note ─────────────────────────────────────────────────────────────
footer_row = kpi_start + 13
ws.row_dimensions[footer_row].height = 20
mset(ws, f'B{footer_row}:E{footer_row}',
     'ℹ️  Data sourced from Crypto Data sheet  |  Coin names: 3–10 chars, no digits',
     bg='ECEFF1', fg='546E7A', italic=True, sz=9)

# ─── Freeze panes ───────────────────────────────────────────────────────────
ws.freeze_panes = 'B8'

# ─── Tab color ──────────────────────────────────────────────────────────────
ws.sheet_properties.tabColor = '1A237E'

# Make data sheet second
data_ws = wb['Crypto Data']
data_ws.sheet_properties.tabColor = '00695C'

# ─── Save ────────────────────────────────────────────────────────────────────
output_path = '/home/claude/Crypto_Comparison_Dashboard.xlsx'
wb.save(output_path)
print(f'Saved: {output_path}')
