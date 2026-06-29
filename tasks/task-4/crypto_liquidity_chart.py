"""
Crypto Liquidity Chart Generator
=================================
- Reads crypto data from crypto_dataset.xlsx
- Filters coins whose names start with vowels (A, E, I, O, U) or B, C, D
- Selects top 10 by Volume (24h) representing liquidity
- Creates an Excel workbook (.xlsm) with a bar chart
- Injects VBA macros for time-based security:
    * Chart visible only between 9 AM and 5 PM
    * Shows warning text outside working hours
    * VBA auto-refreshes every 60 seconds to check time
"""

import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import time

# ─────────────────────────────────────────────────────────────
# STEP 1: Load and filter the data
# ─────────────────────────────────────────────────────────────

INPUT_FILE = r"d:\TASK-4\crypto_dataset.xlsx"
OUTPUT_FILE = r"d:\TASK-4\crypto_liquidity_chart.xlsx"  # Temporary .xlsx
FINAL_FILE = r"d:\TASK-4\Crypto_Top10_Liquidity.xlsm"   # Final .xlsm with VBA

print("=" * 60)
print("  CRYPTO LIQUIDITY CHART GENERATOR")
print("=" * 60)

print("\n[1/5] Loading crypto dataset...")
wb_src = openpyxl.load_workbook(INPUT_FILE)
ws_src = wb_src["Crypto Data"]

# Extract headers
headers = [cell.value for cell in ws_src[1]]
name_idx = headers.index("Coin Name")
volume_idx = headers.index("Volume (24h) (USD)")

# Filter coins starting with vowels or B, C, D
ALLOWED_STARTS = set("AEIOUBCDaeiou bcd")
filtered_coins = []

for row in ws_src.iter_rows(min_row=2, max_row=ws_src.max_row, values_only=True):
    coin_name = row[name_idx]
    volume = row[volume_idx]
    if coin_name and coin_name[0].upper() in ['A', 'E', 'I', 'O', 'U', 'B', 'C', 'D']:
        filtered_coins.append({
            "name": coin_name,
            "volume": volume if volume else 0
        })

# Sort by volume descending and take top 10
filtered_coins.sort(key=lambda x: x["volume"], reverse=True)
top_10 = filtered_coins[:10]

print(f"   Total coins matching filter: {len(filtered_coins)}")
print(f"   Top 10 by Volume (24h):")
for i, coin in enumerate(top_10, 1):
    print(f"     {i:2d}. {coin['name']:<30s} Volume: ${coin['volume']:>20,.0f}")

# ─────────────────────────────────────────────────────────────
# STEP 2: Create the Excel workbook with data
# ─────────────────────────────────────────────────────────────

print("\n[2/5] Creating Excel workbook...")
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Top 10 Liquidity"

# ── Color Palette ──
DARK_BG = "1A1A2E"
HEADER_BG = "16213E"
ACCENT_1 = "0F3460"
ACCENT_2 = "E94560"
GOLD = "FFD700"
WHITE = "FFFFFF"
LIGHT_GRAY = "B0B0B0"
CHART_COLORS = [
    "E94560", "0F3460", "FFD700", "00D9FF", "FF6B35",
    "7B2FF7", "00C49A", "FF4081", "536DFE", "FFAB40"
]

# ── Styling ──
header_font = Font(name="Calibri", bold=True, size=12, color=WHITE)
header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

data_font = Font(name="Calibri", size=11, color=WHITE)
data_fill = PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type="solid")
data_alignment = Alignment(horizontal="center", vertical="center")

alt_fill = PatternFill(start_color="0D1B2A", end_color="0D1B2A", fill_type="solid")
thin_border = Border(
    left=Side(style="thin", color="333355"),
    right=Side(style="thin", color="333355"),
    top=Side(style="thin", color="333355"),
    bottom=Side(style="thin", color="333355"),
)

# ── Title Row ──
ws.merge_cells("A1:C1")
title_cell = ws["A1"]
title_cell.value = "Top 10 Crypto Coins by Liquidity (24h Volume)"
title_cell.font = Font(name="Calibri", bold=True, size=16, color=GOLD)
title_cell.fill = PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type="solid")
title_cell.alignment = Alignment(horizontal="center", vertical="center")
for col in range(1, 4):
    ws.cell(row=1, column=col).fill = PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type="solid")
ws.row_dimensions[1].height = 40

# ── Subtitle Row ──
ws.merge_cells("A2:C2")
subtitle_cell = ws["A2"]
subtitle_cell.value = "Filtered: Coin names starting with Vowels (A,E,I,O,U) or B, C, D"
subtitle_cell.font = Font(name="Calibri", italic=True, size=10, color=LIGHT_GRAY)
subtitle_cell.fill = PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type="solid")
subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
for col in range(1, 4):
    ws.cell(row=2, column=col).fill = PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type="solid")
ws.row_dimensions[2].height = 25

# ── Headers (Row 3) ──
column_headers = ["Rank", "Coin Name", "Volume (24h) USD"]
for col, header in enumerate(column_headers, 1):
    cell = ws.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border
ws.row_dimensions[3].height = 30

# ── Data Rows (Rows 4-13) ──
for i, coin in enumerate(top_10):
    row_num = i + 4
    fill = data_fill if i % 2 == 0 else alt_fill

    # Rank
    rank_cell = ws.cell(row=row_num, column=1, value=i + 1)
    rank_cell.font = Font(name="Calibri", bold=True, size=11, color=GOLD)
    rank_cell.fill = fill
    rank_cell.alignment = data_alignment
    rank_cell.border = thin_border

    # Coin Name
    name_cell = ws.cell(row=row_num, column=2, value=coin["name"])
    name_cell.font = data_font
    name_cell.fill = fill
    name_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    name_cell.border = thin_border

    # Volume
    vol_cell = ws.cell(row=row_num, column=3, value=coin["volume"])
    vol_cell.font = data_font
    vol_cell.fill = fill
    vol_cell.alignment = data_alignment
    vol_cell.number_format = '$#,##0'
    vol_cell.border = thin_border

    ws.row_dimensions[row_num].height = 25

# ── Column Widths ──
ws.column_dimensions["A"].width = 10
ws.column_dimensions["B"].width = 30
ws.column_dimensions["C"].width = 28

# ── Footer ──
footer_row = 15
ws.merge_cells(f"A{footer_row}:C{footer_row}")
footer_cell = ws.cell(row=footer_row, column=1,
                       value="Chart visible only during working hours (9 AM - 5 PM)")
footer_cell.font = Font(name="Calibri", italic=True, size=9, color=ACCENT_2)
footer_cell.fill = PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type="solid")
footer_cell.alignment = Alignment(horizontal="center", vertical="center")
for col in range(1, 4):
    ws.cell(row=footer_row, column=col).fill = PatternFill(
        start_color=DARK_BG, end_color=DARK_BG, fill_type="solid"
    )

# ─────────────────────────────────────────────────────────────
# STEP 3: Create the Bar Chart
# ─────────────────────────────────────────────────────────────

print("\n[3/5] Creating liquidity bar chart...")

chart = BarChart()
chart.type = "col"
chart.style = 10
chart.title = "Top 10 Crypto Coins - Liquidity (24h Volume)"
chart.y_axis.title = "Volume (24h) in USD"
chart.x_axis.title = "Coin Name"
chart.y_axis.numFmt = '$#,##0,,"M"'

# Data reference (Volume values in column C, rows 4-13)
data_ref = Reference(ws, min_col=3, min_row=3, max_row=13)
cats_ref = Reference(ws, min_col=2, min_row=4, max_row=13)

chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.shape = 4

# Style the series with individual colors per bar
series = chart.series[0]
series.graphicalProperties.line.noFill = True

for i, color in enumerate(CHART_COLORS):
    pt = DataPoint(idx=i)
    pt.graphicalProperties.solidFill = color
    series.data_points.append(pt)

# Data labels
series.dLbls = DataLabelList()
series.dLbls.showVal = True
series.dLbls.numFmt = '$#,##0,,"M"'

# Chart size and position
chart.width = 28
chart.height = 18

# Place chart starting at E2
ws.add_chart(chart, "E2")

# ─────────────────────────────────────────────────────────────
# STEP 4: Save as .xlsx first
# ─────────────────────────────────────────────────────────────

print("\n[4/5] Saving temporary workbook...")
wb.save(OUTPUT_FILE)
print(f"   Saved: {OUTPUT_FILE}")

# ─────────────────────────────────────────────────────────────
# STEP 5: Add VBA macros using win32com and save as .xlsm
# ─────────────────────────────────────────────────────────────

print("\n[5/5] Injecting VBA macros for time-based security...")

import win32com.client
import pythoncom

pythoncom.CoInitialize()

try:
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False

    # Open the .xlsx file
    wb_com = excel.Workbooks.Open(os.path.abspath(OUTPUT_FILE))
    ws_com = wb_com.Sheets("Top 10 Liquidity")

    # ── Add a label (text box) for the "outside working hours" message ──
    # We'll add a shape that covers the chart area
    # Position it over the chart area (E2 area, roughly column 5 onwards)
    # Using points: col E starts ~at 250pt, chart is ~750pt wide, ~500pt tall
    left = 290   # roughly column E
    top = 10     # roughly row 2
    width = 750
    height = 480

    shp = ws_com.Shapes.AddTextbox(1, left, top, width, height)  # 1 = msoTextOrientationHorizontal
    shp.Name = "WarningBox"
    shp.TextFrame.Characters().Text = "Please open in working hours ( 9 am to 5 pm )"
    shp.TextFrame.Characters().Font.Size = 28
    shp.TextFrame.Characters().Font.Bold = True
    shp.TextFrame.Characters().Font.Color = 0x004560E9  # Red-ish
    shp.TextFrame.HorizontalAlignment = -4108  # xlCenter
    shp.TextFrame.VerticalAlignment = -4108     # xlCenter
    shp.Fill.ForeColor.RGB = 0x002E1A1A         # Dark background
    shp.Line.ForeColor.RGB = 0x000F3460         # Border color
    shp.Line.Weight = 2
    shp.Visible = False  # Start hidden, VBA will manage

    # ── Inject VBA Code ──
    vba_code = '''
Option Explicit

' ============================================================
' TIME-BASED SECURITY MODULE
' Chart visible only between 9 AM and 5 PM
' Auto-refreshes every 60 seconds
' ============================================================

Dim NextRefreshTime As Date

Sub CheckWorkingHours()
    ' Check if current time is within working hours (9 AM to 5 PM)
    Dim currentHour As Integer
    Dim ws As Worksheet
    Dim chartObj As ChartObject
    Dim warningBox As Shape
    
    currentHour = Hour(Now())
    
    Set ws = ThisWorkbook.Sheets("Top 10 Liquidity")
    
    ' Get the chart object
    If ws.ChartObjects.Count > 0 Then
        Set chartObj = ws.ChartObjects(1)
    End If
    
    ' Get the warning text box
    On Error Resume Next
    Set warningBox = ws.Shapes("WarningBox")
    On Error GoTo 0
    
    If currentHour >= 9 And currentHour < 17 Then
        ' WORKING HOURS: Show chart, hide warning
        If Not chartObj Is Nothing Then
            chartObj.Visible = True
        End If
        If Not warningBox Is Nothing Then
            warningBox.Visible = msoFalse
        End If
    Else
        ' OUTSIDE WORKING HOURS: Hide chart, show warning
        If Not chartObj Is Nothing Then
            chartObj.Visible = False
        End If
        If Not warningBox Is Nothing Then
            warningBox.Visible = msoTrue
        End If
    End If
    
    ' Schedule next refresh in 60 seconds
    NextRefreshTime = Now + TimeSerial(0, 1, 0)
    Application.OnTime NextRefreshTime, "CheckWorkingHours"
End Sub

Sub StopRefresh()
    ' Stop the auto-refresh timer
    On Error Resume Next
    Application.OnTime NextRefreshTime, "CheckWorkingHours", , False
    On Error GoTo 0
End Sub

Sub Auto_Open()
    ' Automatically runs when workbook is opened (for .xls compatibility)
    CheckWorkingHours
End Sub
'''

    # VBA code for ThisWorkbook module (Workbook_Open event)
    workbook_vba = '''
Private Sub Workbook_Open()
    ' Run the time check when workbook opens
    Call CheckWorkingHours
End Sub

Private Sub Workbook_BeforeClose(Cancel As Boolean)
    ' Stop the timer when closing
    Call StopRefresh
End Sub
'''

    # Access VBA Project
    vba_project = wb_com.VBProject

    # Add a standard module for the main code
    vba_module = vba_project.VBComponents.Add(1)  # 1 = vbext_ct_StdModule
    vba_module.Name = "TimeSecurityModule"
    vba_module.CodeModule.AddFromString(vba_code)

    # Add code to ThisWorkbook module
    this_wb = vba_project.VBComponents("ThisWorkbook")
    this_wb.CodeModule.AddFromString(workbook_vba)

    # Save as .xlsm (macro-enabled workbook)
    # FileFormat 52 = xlOpenXMLWorkbookMacroEnabled (.xlsm)
    wb_com.SaveAs(os.path.abspath(FINAL_FILE), FileFormat=52)
    wb_com.Close(SaveChanges=False)
    excel.Quit()

    print(f"   [OK] VBA macros injected successfully!")
    print(f"   [OK] Saved as: {FINAL_FILE}")

except Exception as e:
    print(f"   [ERROR] Error with VBA injection: {e}")
    print(f"   Attempting cleanup...")
    try:
        wb_com.Close(SaveChanges=False)
    except:
        pass
    try:
        excel.Quit()
    except:
        pass
    raise

finally:
    pythoncom.CoUninitialize()

# Clean up temporary .xlsx file
try:
    os.remove(OUTPUT_FILE)
    print(f"   Cleaned up temporary file: {OUTPUT_FILE}")
except:
    pass

print("\n" + "=" * 60)
print("  [OK] DONE! Workbook created successfully.")
print("=" * 60)
print(f"\n  Output: {FINAL_FILE}")
print(f"\n  Features:")
print(f"     - Top 10 crypto coins by 24h Volume (Liquidity)")
print(f"     - Filtered for coins starting with A,E,I,O,U,B,C,D")
print(f"     - Bar chart with individual coin colors")
print(f"     - VBA time-based security:")
print(f"       * Chart visible ONLY from 9 AM to 5 PM")
print(f"       * Warning message shown outside working hours")
print(f"       * Auto-refreshes every 60 seconds")
print(f"\n  NOTE: Enable macros when opening the workbook!")
print("=" * 60)
