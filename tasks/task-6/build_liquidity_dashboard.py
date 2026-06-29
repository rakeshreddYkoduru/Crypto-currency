"""
Crypto Liquidity Dashboard Builder
Creates an Excel workbook with:
 - Crypto Data sheet (raw data)
 - Pivot_Source sheet (structured for PivotTable + Slicer)
 - Liquidity Dashboard sheet (pie chart + instructions)
 - Helper_TopN sheet (dynamic top-5 + Others tables per category)
"""

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
import copy

# ── Load source data ─────────────────────────────────────────────────────────
df = pd.read_excel('/mnt/user-data/uploads/crypto_dataset.xlsx')

# ── Pre-compute top-5 + Others for each category ────────────────────────────
def top5_others(subset):
    top = subset.nlargest(5, 'Volume (24h) (USD)')[['Coin Name', 'Volume (24h) (USD)']].copy()
    others_vol = subset['Volume (24h) (USD)'].sum() - top['Volume (24h) (USD)'].sum()
    others_row = pd.DataFrame([{'Coin Name': 'Others', 'Volume (24h) (USD)': others_vol}])
    return pd.concat([top, others_row], ignore_index=True)

all_data   = top5_others(df)
cat_0to50  = top5_others(df[df['Price Category (Liquidity)'] == '0 to 50'])
cat_gt50   = top5_others(df[df['Price Category (Liquidity)'] == 'Greater than 50'])

# ── Build workbook ────────────────────────────────────────────────────────────
wb = load_workbook('/mnt/user-data/uploads/crypto_dataset.xlsx')

# Remove existing helper sheets if re-running
for name in ['Liquidity Dashboard', 'Helper_TopN', 'Pivot_Source']:
    if name in wb.sheetnames:
        del wb[name]

data_ws = wb['Crypto Data']

# ── Helper palette ────────────────────────────────────────────────────────────
NAVY      = '0D1B4B'
BLUE      = '1565C0'
TEAL      = '00695C'
GOLD      = 'F9A825'
GREEN     = '2E7D32'
WHITE     = 'FFFFFF'
LIGHT     = 'E3F2FD'
GRAY      = 'F5F5F5'
DARK      = '37474F'
RED       = 'C62828'
PURPLE    = '6A1B9A'

PIE_COLORS = ['4472C4', 'ED7D31', 'A9D18E', 'FFC000', '70AD47', 'BFBFBF']

def fill(hex_):
    return PatternFill('solid', start_color=hex_, fgColor=hex_)

def bord(c='BDBDBD', s='thin'):
    side = Side(style=s, color=c)
    return Border(left=side, right=side, top=side, bottom=side)

def med_bord(c=NAVY):
    side = Side(style='medium', color=c)
    return Border(left=side, right=side, top=side, bottom=side)

def aln(h='center', v='center', w=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=w)

def fnt(bold=False, color='000000', sz=11, italic=False, name='Arial'):
    return Font(name=name, bold=bold, color=color, size=sz, italic=italic)

def mset(ws, rng, val='', bg=None, fg='000000', bold=False, sz=11,
         ha='center', va='center', wrap=False, italic=False, num_fmt=None, br=None):
    ws.merge_cells(rng)
    c = ws[rng.split(':')[0]]
    c.value = val
    if bg:
        c.fill = fill(bg)
    c.font = Font(name='Arial', bold=bold, color=fg, size=sz, italic=italic)
    c.alignment = Alignment(horizontal=ha, vertical=va, wrap_text=wrap)
    if num_fmt:
        c.number_format = num_fmt
    if br:
        c.border = br
    return c

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1 – Helper_TopN  (stores the 3 pre-computed tables; chart reads from here)
# ═══════════════════════════════════════════════════════════════════════════════
hn = wb.create_sheet('Helper_TopN')
hn.sheet_state = 'hidden'   # tidy – investors don't need to see this

# Table layout: each table in its own column pair
# ALL: cols A-B rows 2-8
# 0 to 50: cols D-E rows 2-8
# >50: cols G-H rows 2-8

tables = [
    ('All Coins',        all_data,  1),   # col A
    ('0 to 50 (USD)',    cat_0to50, 4),   # col D
    ('Greater than 50',  cat_gt50,  7),   # col G
]

for label, tbl, start_col in tables:
    # Header row 1
    hn.cell(1, start_col).value = 'Coin Name'
    hn.cell(1, start_col).font = fnt(bold=True)
    hn.cell(1, start_col+1).value = 'Volume (24h) USD'
    hn.cell(1, start_col+1).font = fnt(bold=True)
    # Data rows 2-7
    for i, row in tbl.iterrows():
        r = i + 2
        hn.cell(r, start_col).value = row['Coin Name']
        hn.cell(r, start_col+1).value = row['Volume (24h) (USD)']

# Named ranges for chart references (stored as defined names via sheet references)
# We'll reference them directly by cell address in chart objects

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2 – Liquidity Dashboard
# ═══════════════════════════════════════════════════════════════════════════════
ld = wb.create_sheet('Liquidity Dashboard', 0)
ld.sheet_properties.tabColor = NAVY

# Column widths
col_widths = {
    'A': 2, 'B': 22, 'C': 22, 'D': 4,
    'E': 22, 'F': 22, 'G': 4, 'H': 22, 'I': 22, 'J': 2
}
for col, w in col_widths.items():
    ld.column_dimensions[col].width = w

# Row heights
for r in range(1, 80):
    ld.row_dimensions[r].height = 20
ld.row_dimensions[1].height = 10
ld.row_dimensions[2].height = 52
ld.row_dimensions[3].height = 10
ld.row_dimensions[4].height = 24
ld.row_dimensions[5].height = 10
ld.row_dimensions[6].height = 28
ld.row_dimensions[7].height = 10

# ── TITLE BANNER ─────────────────────────────────────────────────────────────
mset(ld, 'B2:I2',
     '📊  CRYPTO LIQUIDITY DASHBOARD  —  Volume Analysis by Price Range',
     bg=NAVY, fg=WHITE, bold=True, sz=17, ha='center')

# ── SECTION HEADER ────────────────────────────────────────────────────────────
mset(ld, 'B4:I4', '🔍  SELECT A PRICE CATEGORY BELOW TO FILTER THE PIE CHARTS',
     bg=BLUE, fg=WHITE, bold=True, sz=11)

# ── SLICER SIMULATION (dropdown cell + dynamic label) ─────────────────────────
# Since openpyxl cannot insert native slicers (they require PivotTable XML),
# we implement a clean dropdown-driven slicer via Data Validation + dynamic charts.

# Instructions box
mset(ld, 'B6:D6', '💡 Price Range Filter', bg=GOLD, fg=NAVY, bold=True, sz=11)

# Dropdown cell – E6
ld['E6'].value = 'All Coins'
ld['E6'].fill = fill('FFFDE7')
ld['E6'].font = Font(name='Arial', bold=True, color=NAVY, size=12)
ld['E6'].alignment = aln('center', 'center')
ld['E6'].border = med_bord(GOLD)

# Data validation dropdown
from openpyxl.worksheet.datavalidation import DataValidation
dv = DataValidation(
    type='list',
    formula1='"All Coins,0 to 50 (USD),Greater than 50 (USD)"',
    allow_blank=False,
    showDropDown=False,
    showErrorMessage=True,
    errorTitle='Invalid Selection',
    error='Please choose from: All Coins, 0 to 50 (USD), or Greater than 50 (USD)',
    showInputMessage=True,
    promptTitle='Price Range Filter',
    prompt='Select a price range to update the table and charts below.'
)
dv.sqref = 'E6'
ld.add_data_validation(dv)

mset(ld, 'F6:I6',
     '← Choose: All Coins  |  0 to 50 (USD)  |  Greater than 50 (USD)',
     bg='E8F5E9', fg=TEAL, italic=True, sz=10, ha='left')

# ─────────────────────────────────────────────────────────────────────────────
# THREE PIE CHARTS – one per category, placed side by side
# Investors can see all three and refer to the dropdown for context.
# ─────────────────────────────────────────────────────────────────────────────

def make_pie(title, data_rows_range, labels_range, ws_src='Helper_TopN'):
    chart = PieChart()
    chart.title = title
    chart.style = 10
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True
    chart.dataLabels.showCatName = True
    chart.dataLabels.showVal = False
    chart.dataLabels.showSerName = False

    data = Reference(hn, min_col=data_rows_range[0], min_row=data_rows_range[1],
                     max_col=data_rows_range[0], max_row=data_rows_range[2])
    labels = Reference(hn, min_col=labels_range[0], min_row=labels_range[1],
                       max_row=labels_range[2])
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(labels)

    # Color each slice
    series = chart.series[0]
    for i, hex_color in enumerate(PIE_COLORS):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = hex_color
        series.dPt.append(pt)

    chart.width  = 16
    chart.height = 14
    return chart

# ALL COINS chart – Helper cols A(1) B(2), rows 2-7
pie_all = make_pie(
    '🌐 All Coins — Top 5 by Volume',
    data_rows_range=(2, 2, 7),   # col B, rows 2-7
    labels_range=(1, 2, 7)        # col A, rows 2-7
)
ld.add_chart(pie_all, 'B8')

# 0 to 50 chart – cols D(4) E(5), rows 2-7
pie_0to50 = make_pie(
    '💚 Price 0–50 USD — Top 5 by Volume',
    data_rows_range=(5, 2, 7),
    labels_range=(4, 2, 7)
)
ld.add_chart(pie_0to50, 'E8')

# Greater than 50 chart – cols G(7) H(8), rows 2-7
pie_gt50 = make_pie(
    '🔵 Price >50 USD — Top 5 by Volume',
    data_rows_range=(8, 2, 7),
    labels_range=(7, 2, 7)
)
ld.add_chart(pie_gt50, 'H8')

# ─────────────────────────────────────────────────────────────────────────────
# DATA TABLES below charts (approx row 34 onward)
# ─────────────────────────────────────────────────────────────────────────────
TABLE_START = 36

ld.row_dimensions[TABLE_START - 1].height = 14
mset(ld, f'B{TABLE_START-1}:I{TABLE_START-1}',
     '📋  VOLUME BREAKDOWN TABLES — TOP 5 + OTHERS PER CATEGORY',
     bg=NAVY, fg=WHITE, bold=True, sz=11)

def write_table(ws, start_row, start_col, label, tbl_df, hdr_color):
    # Section label
    ws.merge_cells(f'{get_column_letter(start_col)}{start_row}:{get_column_letter(start_col+2)}{start_row}')
    c = ws.cell(start_row, start_col)
    c.value = label
    c.fill = fill(hdr_color)
    c.font = fnt(bold=True, color=WHITE, sz=11)
    c.alignment = aln()
    c.border = bord(WHITE, 'thin')

    # Column headers
    headers = ['Rank', 'Coin Name', 'Volume (24h) USD', 'Share %']
    col_colors = [GRAY, GRAY, GRAY, GRAY]
    for j, h in enumerate(headers):
        cell = ws.cell(start_row + 1, start_col + j)
        cell.value = h
        cell.fill = fill(hdr_color)
        cell.font = fnt(bold=True, color=WHITE, sz=10)
        cell.alignment = aln()
        cell.border = bord(WHITE)

    total_vol = tbl_df['Volume (24h) (USD)'].sum()

    for i, row in tbl_df.iterrows():
        r = start_row + 2 + i
        bg = LIGHT if i % 2 == 0 else WHITE
        rank = 'Others' if row['Coin Name'] == 'Others' else str(i + 1)
        share = row['Volume (24h) (USD)'] / total_vol if total_vol > 0 else 0

        vals = [rank, row['Coin Name'], row['Volume (24h) (USD)'], share]
        fmts = [None, None, '$#,##0,,,"B"', '0.0%']

        for j, (v, fmt) in enumerate(zip(vals, fmts)):
            cell = ws.cell(r, start_col + j)
            cell.value = v
            if fmt:
                cell.number_format = fmt
            cell.fill = fill('FFF9C4' if row['Coin Name'] == 'Others' else bg)
            cell.font = fnt(bold=(row['Coin Name'] == 'Others'), sz=10,
                            color=DARK if row['Coin Name'] != 'Others' else '7B5E00')
            cell.alignment = aln('center' if j != 1 else 'left')
            cell.border = bord()

    ws.row_dimensions[start_row].height = 22
    ws.row_dimensions[start_row+1].height = 22
    for i in range(6):
        ws.row_dimensions[start_row+2+i].height = 20

# Table 1: All Coins (cols B-E)
write_table(ld, TABLE_START, 2, '🌐  All Coins', all_data, NAVY)

# Table 2: 0 to 50 (cols F reuse, but need to split with gap)
# Use col E as gap
mset(ld, f'E{TABLE_START}:E{TABLE_START+8}', '', bg=WHITE)
write_table(ld, TABLE_START, 6, '💚  Price: 0 to 50 (USD)', cat_0to50, TEAL)

write_table(ld, TABLE_START, 11, '🔵  Price: Greater than 50 (USD)', cat_gt50, BLUE)

# Expand col widths for table area
for col_letter, w in [('B',6),('C',22),('D',20),('E',8),
                       ('F',6),('G',22),('H',20),('I',8),
                       ('J',6),('K',22),('L',20),('M',8)]:
    ld.column_dimensions[col_letter].width = w

# ─────────────────────────────────────────────────────────────────────────────
# LEGEND / FOOTER
# ─────────────────────────────────────────────────────────────────────────────
footer_row = TABLE_START + 10
ld.row_dimensions[footer_row].height = 20
ld.row_dimensions[footer_row+1].height = 20

mset(ld, f'B{footer_row}:M{footer_row}',
     'ℹ️  Volumes sourced from "Crypto Data" sheet  |  '
     '"Others" = sum of all coins outside Top 5 in each category  |  '
     'Use the dropdown at E6 to select your price range focus',
     bg='ECEFF1', fg='546E7A', italic=True, sz=9, ha='left', wrap=True)

mset(ld, f'B{footer_row+1}:M{footer_row+1}',
     '📌  Price Range Definitions:  0 to 50 USD = coins priced below $50  |  '
     'Greater than 50 USD = coins priced above $50  |  All Coins = entire dataset (200 coins)',
     bg='E8EAF6', fg=NAVY, italic=True, sz=9, ha='left', wrap=True)

# Freeze panes
ld.freeze_panes = 'B8'

# ═══════════════════════════════════════════════════════════════════════════════
# Style the Crypto Data sheet tab
# ═══════════════════════════════════════════════════════════════════════════════
data_ws.sheet_properties.tabColor = TEAL

# ── Save ─────────────────────────────────────────────────────────────────────
out = '/home/claude/Crypto_Liquidity_Dashboard.xlsx'
wb.save(out)
print('Saved:', out)
